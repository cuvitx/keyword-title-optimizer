import streamlit as st
import pandas as pd
import spacy
import time
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

st.set_page_config(
    page_title="Keywords Title Matcher",
    initial_sidebar_state="collapsed",
    page_icon="📈"
)
st.markdown("<style> .e1fqkh3o1 {display: none;} </style>", unsafe_allow_html=True)

st.markdown("# Keywords Title Matcher")

nlp = spacy.load("sv_core_news_lg")

with st.form(key='2_form'):
    uploaded_file = st.file_uploader("Choose the Excel file:")
    Keyword_sheet = st.text_input("Enter the Keywords Sheet Name:")
    submit_button = st.form_submit_button(label='Proceed')
    if submit_button:
        with pd.ExcelFile(uploaded_file) as xls:
            df_keywords = pd.read_excel(xls, sheet_name=Keyword_sheet, usecols=['Keyword', 'Volume'])
            df_keywords.dropna(inplace=True)
            keywords = list(df_keywords['Keyword'])

            tabs = xls.sheet_names
            total_titles = sum([1 for u in tabs if u != 'Generic KW']) * 200
            progress_bar = st.progress(0)
            processed_titles = 0
            elapsed_time_list = []

            status_text = st.empty()

            # Create a new Excel workbook to store the results
            output_workbook = Workbook()
            output_workbook.remove(output_workbook.active)  # Remove the default sheet

            # Liste pour stocker tous les titres de tous les onglets
            all_titles = []

            # Dictionnaire pour stocker les meilleurs mots clés correspondants pour chaque titre
            best_keywords = {}

            for u in tabs:
                if u != 'Generic KW':
                    df_titles = pd.read_excel(xls, sheet_name=u, usecols=['Title'])
                    titles = list(df_titles['Title'])
                    all_titles.extend(titles)

            # Traiter tous les titres
            for title in nlp.pipe(all_titles):
                start_time = time.time()
                keyword_docs = list(nlp.pipe(keywords))
                keyword_hashes = [doc.similarity(title) for doc in keyword_docs]
                best_match = max(keyword_docs, key=lambda keyword: title.similarity(keyword) if title.similarity(
                                         keyword) > 0.5 else -1)
                if best_match:
                    best_keywords[title.text] = best_match.text
                    keywords.remove(best_match.text)  # Remove the matched keyword from the list

                elapsed_time = time.time() - start_time
                elapsed_time_list.append(elapsed_time)

                if len(elapsed_time_list) > 50:
                    elapsed_time_list.pop(0)

                processed_titles += 1
                progress = processed_titles / total_titles
                progress_bar.progress(progress)

                if processed_titles < total_titles and len(elapsed_time_list) > 0:
                    avg_elapsed_time = sum(elapsed_time_list) / len(elapsed_time_list)
                    remaining_time = avg_elapsed_time * (total_titles - processed_titles)
                    status_text.text(f"Estimated time remaining : {int(remaining_time // 60)} minutes {int(remaining_time % 60)} seconds")
                else:
                    status_text.text("Processing complete!")

            # Attribuer les meilleurs mots clés correspondants aux titres dans chaque onglet
            for u in tabs:
                if u != 'Generic KW':
                    df_titles = pd.read_excel(xls, sheet_name=u, usecols=['Title'])
                    matched_keywords = [best_keywords.get(title) for title in df_titles['Title']]
                    df_titles['Matched_Keyword'] = matched_keywords

                    # Add the modified DataFrame as a new sheet in the output workbook
                    output_sheet = output_workbook.create_sheet(u)
                    for r in dataframe_to_rows(df_titles, index=False, header=True):
                        output_sheet.append(r)

            # Add the keywords sheet with used keywords in blue and the "Volume de recherche" column
            used_keywords = list(set(df_keywords['Keyword']) - set(keywords))
            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            keyword_sheet = output_workbook.create_sheet(Keyword_sheet)
            for r in dataframe_to_rows(df_keywords, index=False, header=True):
                keyword_sheet.append(r)

            for row in keyword_sheet.iter_rows():
                for cell in row:
                    if cell.row > 1 and cell.column == 1:  # Check if it's a keyword cell in the data rows
                        if cell.value in used_keywords:
                            cell.fill = blue_fill  # Fill the used keywords with blue color

            # Save the output workbook with all the sheets
            output_workbook.save('Matched_Keywords.xlsx')


